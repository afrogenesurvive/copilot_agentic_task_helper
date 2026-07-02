#!/usr/bin/env python3
"""
Convert monthly Bills Check .numbers files to .xlsx, preserving ALL rows
including the '$' row (row 5) that was previously excluded from the master.

For each .numbers file in ~/Downloads, creates a corresponding .xlsx in
~/Downloads/fixed_xlsx/ with every row included.

Usage:
    python3 convert_numbers_to_xlsx.py
"""

import os, glob, re
from numbers_parser import Document
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

HOME = os.path.expanduser("~")
SRC_DIR = os.path.join(HOME, "Downloads")
OUT_DIR = os.path.join(HOME, "Downloads", "fixed_xlsx")

# Ensure output directory exists
os.makedirs(OUT_DIR, exist_ok=True)

# Find all Bills Check .numbers files
pattern = os.path.join(SRC_DIR, "Bills Check *.numbers")
files = sorted(glob.glob(pattern))

print(f"Found {len(files)} .numbers files to convert\n")

for fpath in files:
    fname = os.path.basename(fpath)
    
    # Parse the numbers file
    doc = Document(fpath)
    sheet = doc.sheets[0]
    table = sheet.tables[0]
    num_rows = table.num_rows
    num_cols = table.num_cols
    
    # Build the output filename
    out_name = re.sub(r'\.numbers$', '.xlsx', fname)
    out_path = os.path.join(OUT_DIR, out_name)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet 1'
    
    # Track where $ rows occur
    dollar_rows = 0
    
    # Copy all rows
    copied_count = 0
    for r in range(num_rows):
        row_has_data = False
        for c in range(num_cols):
            try:
                val = table.cell(r, c).value
            except:
                val = None
            
            if val is not None:
                row_has_data = True
            
            # Set value in xlsx cell (1-indexed rows and cols)
            ws.cell(row=r+1, column=c+1).value = val
        
        if row_has_data:
            copied_count += 1
            # Check if this is a $ row
            a = table.cell(r, 0).value
            if a is not None and str(a).strip() == '$':
                dollar_rows += 1
    
    wb.save(out_path)
    dollar_info = f" (including {dollar_rows} '$' row{'s' if dollar_rows != 1 else ''})" if dollar_rows else ""
    print(f"  ✓ {out_name} — {copied_count} rows, {num_cols} cols{dollar_info}")

print(f"\n✅ All files saved to: {OUT_DIR}")
print(f"   Total: {len(files)} files")
