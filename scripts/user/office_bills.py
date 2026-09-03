#!/usr/bin/env python3
"""
office_bills.py — Monthly office bills extraction (standalone).

A direct port of docs/safe/office_bills.prompt.md to an executable script.
It does NOT rely on the Gmail/Drive MCP servers — it talks to the Gmail and
Google Drive REST APIs directly using the OAuth tokens already on disk:

  • Drive writes use safe/tokens/gmail-token.json + safe/gmail-oauth2.json
    (the Drive-scoped token — NOT the .env token).
  • Bill emails are read from the entclinicmobay inbox using the refresh token
    in the repo .env (GMAIL_REFRESH_TOKEN_2 / GMAIL_USER_2) + GMAIL_CLIENT_ID/SECRET.

Workflow (mirrors the prompt):
  1. Download the current "Bills Check Master" xlsx from Drive.
  2. Find the most recent month header → previous-month block is the layout
     template (vendor headers, subheaders, blank rows, col G False).
  3. Search the entclinicmobay inbox for the target month's bill emails.
  4. Extract amounts / account numbers / due dates.
  5. Clone the previous month block to the top, change the month label, and
     overwrite amount/due on the rows whose account numbers matched email.
     Vendors with no email (CC CABLE, GARBAGE, COLUMBUS, FLOW JAMAICA internet)
     are carried forward by the clone automatically.
  6. Validate + upload back to Drive (PATCH media) — unless --dry-run.

Usage:
    python3 scripts/user/office_bills.py                  # real run (uploads)
    python3 scripts/user/office_bills.py --dry-run        # build locally only
    python3 scripts/user/office_bills.py --month 2026-09  # pick the target month
    python3 scripts/user/office_bills.py --inspect        # read-only layout dump

Never run the real upload twice for the same month — an idempotency guard
aborts if the target month header is already present (override with --force).
"""

import argparse
import base64
import calendar
import copy
import json
import os
import re
import sys
import urllib.parse
import urllib.request

import openpyxl

# ───────────────────────── Config ─────────────────────────
DEFAULT_FILE_ID = "1ssq9svNbAy2wBEUfwVzL7KfhwEkEwJYl"  # Bills Check Master.xlsx
VENDOR_ORDER = ["FLOW", "JPS", "NWC", "CC CABLE", "GARBAGE", "COLUMBUS", "FLOW JAMAICA", "HAWKEYE"]
NO_EMAIL_VENDORS = {"CC CABLE", "GARBAGE", "COLUMBUS"}  # FLOW JAMAICA partially emailed (47546801)

# Gmail senders → where the values live. Queries match docs/safe/office_bills.prompt.md.
SENDERS = {
    "FLOW":    {"query": "from:Flow-billpay-alert@cwc.com", "where": "html", "max": 10},
    "JPS":     {"query": "from:no-reply@jpsco.com",         "where": "subject", "max": 10},
    "NWC":     {"query": "from:ebill@ebill.nwc.com.jm",     "where": "subject", "max": 10},
    "HAWKEYE": {"query": "from:hawkeye",                    "where": "subject", "max": 5},
}

# FLOW email account number → sheet ACCOUNT cell value (verified mapping from the prompt).
FLOW_ACCT_MAP = {
    "870022420000": "87002242/00002",   # FLOW row 3 (phone 952-4099/2242)
    "870032650000": "96929857/00009",   # FLOW row 4 (usually blank / credit)
    "50741183":     "50741183",         # FLOW row 5 — New Office Line(s), NOT FLOW JAMAICA
    "47546801":     "47546801",         # FLOW JAMAICA USD row
}
# JPS fixed spreadsheet order (DO NOT reorder) + NWC order (1st row = n/a).
JPS_ACCTS = ["437336-439977", "437336-878676", "437336-914923", "437336-440015"]
NWC_ACCTS = ["1209548-1209538", "1421964-1125635", "1316541-1316531"]

MONTH_RE = re.compile(
    r"^(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})$"
)

# Cell layout used across the master (see safe/reorder_and_format.py):
#   col B = PHONE #/CIS# / ADDRESS / vendor name
#   col C = ACCOUNT
#   col D = AMOUNT DUE
#   col E = DUE DATE
#   col F = AMOUNT PAID
#   col G = COMPLETE (False for every data row)
COL_ACCOUNT, COL_AMOUNT, COL_DUE, COL_COMPLETE = 3, 4, 5, 7


# ───────────────────────── Paths / creds ─────────────────────────
def repo_root():
    # <repo>/scripts/user/office_bills.py → up three dirnames = <repo>
    return os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def load_env(env_path):
    out = {}
    try:
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                m = re.match(r"^([A-Za-z0-9_]+)=(.*)$", line.strip())
                if m:
                    out[m.group(1)] = m.group(2).strip()
    except FileNotFoundError:
        pass
    return out


# ───────────────────────── OAuth ─────────────────────────
def refresh_oauth(client_id, client_secret, refresh_token):
    """Return a fresh access token for the given OAuth client + refresh token."""
    body = urllib.parse.urlencode({
        "client_id": client_id,
        "client_secret": client_secret,
        "refresh_token": refresh_token,
        "grant_type": "refresh_token",
    }).encode()
    req = urllib.request.Request("https://oauth2.googleapis.com/token", data=body, method="POST")
    with urllib.request.urlopen(req) as resp:
        data = json.loads(resp.read())
    return data["access_token"]


def drive_token():
    """Refresh + return the Drive-scoped access token (safe/tokens/gmail-token.json)."""
    repo = repo_root()
    creds_path = os.path.join(repo, "safe", "gmail-oauth2.json")
    token_path = os.path.join(repo, "safe", "tokens", "gmail-token.json")
    with open(creds_path) as f:
        creds = json.load(f)
    with open(token_path) as f:
        token = json.load(f)
    installed = creds.get("installed", creds)
    access = refresh_oauth(installed["client_id"], installed["client_secret"], token["refresh_token"])
    token["access_token"] = access
    with open(token_path, "w") as f:
        json.dump(token, f, indent=2)
    return access


def gmail_token():
    """Return (access_token, userId) for the entclinicmobay inbox from repo .env."""
    env = load_env(os.path.join(repo_root(), ".env"))
    cid = env.get("GMAIL_CLIENT_ID")
    secret = env.get("GMAIL_CLIENT_SECRET")
    refresh = env.get("GMAIL_REFRESH_TOKEN_2")
    user = env.get("GMAIL_USER_2") or "entclinicmobay@gmail.com"
    if not (cid and secret and refresh):
        raise SystemExit("Missing GMAIL_CLIENT_ID/SECRET or GMAIL_REFRESH_TOKEN_2 in .env (entclinicmobay).")
    return refresh_oauth(cid, secret, refresh), user


def drive_request(access, path, method="GET", body=None, headers=None, raw_body=None, content_type=None):
    url = "https://www.googleapis.com" + path
    h = {"Authorization": f"Bearer {access}"}
    if headers:
        h.update(headers)
    data = None
    if raw_body is not None:
        data = raw_body
        h["Content-Type"] = content_type or "application/octet-stream"
    elif body is not None:
        data = json.dumps(body).encode()
        h["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=data, method=method, headers=h)
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())


def download_master(file_id, access, dest):
    url = f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {access}"})
    with urllib.request.urlopen(req) as resp:
        with open(dest, "wb") as f:
            f.write(resp.read())
    return dest


def upload_master(file_id, access, src):
    info = drive_request(access, f"/drive/v3/files/{file_id}?fields=mimeType")
    with open(src, "rb") as f:
        content = f.read()
    url = f"https://www.googleapis.com/upload/drive/v3/files/{file_id}?uploadType=media"
    req = urllib.request.Request(url, data=content, method="PATCH")
    req.add_header("Authorization", f"Bearer {access}")
    req.add_header("Content-Type", info.get("mimeType", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
    req.add_header("Content-Length", str(len(content)))
    with urllib.request.urlopen(req) as resp:
        result = json.loads(resp.read())
    return result.get("name")


# ───────────────────────── Gmail ─────────────────────────
def gmail_request(access, path, params=None):
    url = "https://gmail.googleapis.com/gmail/v1/" + path.lstrip("/")
    if params:
        url += "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {access}"})
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())


def gmail_list(access, user, query, max_results):
    return gmail_request(access, f"users/{user}/messages",
                         {"q": query, "maxResults": max_results}).get("messages", [])


def _decode(data):
    pad = data + "=" * (-len(data) % 4)
    return base64.urlsafe_b64decode(pad).decode("utf-8", "replace")


def _walk_parts(part, acc):
    mime = part.get("mimeType", "")
    if part.get("body", {}).get("data"):
        acc.setdefault(mime, []).append(_decode(part["body"]["data"]))
    for sub in part.get("parts", []):
        _walk_parts(sub, acc)


def gmail_get(access, user, msg_id):
    msg = gmail_request(access, f"users/{user}/messages/{msg_id}", {"format": "full"})
    payload = msg.get("payload", {})
    acc = {}
    _walk_parts(payload, acc)
    html = "\n".join(acc.get("text/html", []))
    text = "\n".join(acc.get("text/plain", []))
    return {
        "id": msg.get("id"),
        "subject": next((h["value"] for h in payload.get("headers", []) if h.get("name") == "Subject"), ""),
        "html": html,
        "text": text,
    }


def parse_amount(raw):
    if raw is None:
        return None
    s = str(raw).strip().replace(",", "")
    neg = s.endswith("-")
    s = s.rstrip("-")
    m = re.match(r"^[^\d]*(-?\d+(?:\.\d+)?)", s)
    if not m:
        return None
    val = float(m.group(1).replace(",", ""))
    return -val if neg else val


# ───────────────────────── Extraction ─────────────────────────
def extract_flow(msg):
    """FLOW sends an HTML body — strip <style> blocks first, then tags."""
    html = msg.get("html") or msg.get("text") or ""
    text = re.sub(r"<style[^>]*>[\s\S]*?</style>", "", html, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"&nbsp;", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    acct = re.search(r"Account Number\s*:?\s*(\d+)", text, re.IGNORECASE)
    amt = re.search(r"Amount Due\s*:?\s*\$?\s*([\d.,\-]+)", text, re.IGNORECASE)
    due = re.search(r"Due Date\s*:?\s*([\dA-Za-z\-]+)", text, re.IGNORECASE)
    return {
        "account": acct.group(1) if acct else None,
        "amount": parse_amount(amt.group(1)) if amt else None,
        "due_date": due.group(1) if due else None,
    }


def extract_subject(msg):
    """JPS / NWC / Hawkeye put the values in the subject line."""
    s = msg.get("subject") or ""
    acct = re.search(r"A/C#:\s*([\d\-]+)", s, re.IGNORECASE)
    amt = re.search(r"Amount:\s*\$?\s*([\d.,\-]+)", s, re.IGNORECASE)
    due = re.search(r"Due Date:\s*([\w\-]+)", s, re.IGNORECASE)
    return {
        "account": acct.group(1) if acct else None,
        "amount": parse_amount(amt.group(1)) if amt else None,
        "due_date": due.group(1) if due else None,
    }


def fetch_bills(access, user, target_ym):
    """Query Gmail for each vendor and return {vendor: [bill, ...]} (latest per account)."""
    q_date = target_ym + "/01"
    found = {v: {} for v in SENDERS}
    for vendor, cfg in SENDERS.items():
        query = f"{cfg['query']} after:{q_date}"
        try:
            msgs = gmail_list(access, user, query, cfg["max"])
        except Exception as e:
            print(f"  ⚠️  Gmail query failed for {vendor}: {e}")
            continue
        extractor = extract_flow if cfg["where"] == "html" else extract_subject
        for m in msgs:
            try:
                msg = gmail_get(access, user, m["id"])
                bill = extractor(msg)
            except Exception as e:
                print(f"  ⚠️  Could not read message {m.get('id')} for {vendor}: {e}")
                continue
            if bill.get("account"):
                found[vendor][bill["account"]] = bill  # keep latest per account
        print(f"  {vendor}: {len(found[vendor])} email(s) matched")
    return {v: list(d.values()) for v, d in found.items()}


# ───────────────────────── Workbook helpers ─────────────────────────
def copy_cell(src, dst):
    dst.value = copy.copy(src.value)
    if src.has_style:
        dst.font = copy.copy(src.font)
        dst.border = copy.copy(src.border)
        dst.fill = copy.copy(src.fill)
        dst.number_format = copy.copy(src.number_format)
        dst.protection = copy.copy(src.protection)
        dst.alignment = copy.copy(src.alignment)


def find_month_headers(ws):
    """Return [(row, label, (year, month))] for every 'Month YYYY' cell in col A."""
    headers = []
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        m = MONTH_RE.match(str(val).strip()) if val is not None else None
        if m:
            mon = list(calendar.month_name).index(m.group(1))
            headers.append((r, str(val).strip(), (int(m.group(2)), mon)))
    return headers


def clone_rows(ws, src_start, src_end, dest_start, row_heights=True):
    """Copy values + styles from src rows to dest rows (same column count)."""
    for off in range(src_end - src_start + 1):
        s, d = src_start + off, dest_start + off
        for c in range(1, ws.max_column + 1):
            copy_cell(ws.cell(s, c), ws.cell(d, c))
        if row_heights and s in ws.row_dimensions:
            ws.row_dimensions[d].height = ws.row_dimensions[s].height


def normalize(token):
    return str(token or "").replace(" ", "").replace("/", "").lower()


def acct_matches(cell_val, token):
    """Tolerant account match: exact match on the normalized (no space/slash) value."""
    cv = normalize(cell_val)
    tk = normalize(token)
    if not cv or not tk:
        return False
    return cv == tk


def overlay_bills(ws, bills, r0, r1, updates):
    """For each bill, find its row in the clone (by account) and set D/E."""
    # Build an index: account-cell (col C, fallback col B) -> row, within r0..r1
    index = []
    for r in range(r0, r1 + 1):
        c = str(ws.cell(r, COL_ACCOUNT).value or "").strip()
        b = str(ws.cell(r, 2).value or "").strip()
        if c or b:
            index.append((r, c, b))

    for vendor, vlist in bills.items():
        for bill in vlist:
            acct = bill.get("account")
            if not acct:
                continue
            # Resolve email account → sheet account token(s)
            if vendor == "FLOW" and acct in FLOW_ACCT_MAP:
                tokens = [FLOW_ACCT_MAP[acct]]
                section = "FLOW JAMAICA" if acct == "47546801" else "FLOW"
            else:
                tokens = [acct]
                section = vendor
            row = None
            for r, c, b in index:
                if any(acct_matches(c, t) or (b and acct_matches(b, t)) for t in tokens):
                    row = r
                    break
            if row is None:
                updates.append(f"  ⚠️  {vendor} acct {acct}: no matching row found (carried forward)")
                continue
            if bill.get("amount") is not None:
                ws.cell(row, COL_AMOUNT).value = bill["amount"]
                updates.append(f"  ✓ {vendor} {acct}: amount → {bill['amount']}")
            if bill.get("due_date"):
                ws.cell(row, COL_DUE).value = str(bill["due_date"])
                updates.append(f"  ✓ {vendor} {acct}: due → {bill['due_date']}")
            # COMPLETE must stay False
            ws.cell(row, COL_COMPLETE).value = False


# ───────────────────────── Inspect ─────────────────────────
def cmd_inspect(file_id, access, tmp):
    download_master(file_id, access, tmp)
    wb = openpyxl.load_workbook(tmp, data_only=False)
    ws = wb.active
    print(f"Sheet: {ws.title}  dims: {ws.max_row} rows x {ws.max_column} cols")
    print("Month headers (col A):")
    for r, label, _ in find_month_headers(ws):
        print(f"  row {r}: {label}")
    print("\nFirst 60 rows (cols A..G):")
    for r in range(1, min(ws.max_row, 60) + 1):
        vals = [ws.cell(r, c).value for c in range(1, 8)]
        if all(v is None for v in vals):
            continue
        shown = []
        for v in vals:
            s = "" if v is None else str(v)
            shown.append(s if len(s) <= 26 else s[:23] + "...")
        print(f"{r:>3} | " + " | ".join(shown))
    return 0


# ───────────────────────── Main ─────────────────────────
def main():
    ap = argparse.ArgumentParser(description="Monthly office bills extraction (xlsx route).")
    ap.add_argument("--month", help="Target month as YYYY-MM (default: month after the last header in the master)")
    ap.add_argument("--file-id", default=DEFAULT_FILE_ID, help="Drive file id of the master xlsx")
    ap.add_argument("--dry-run", action="store_true", help="Download + build + validate, but do NOT upload")
    ap.add_argument("--inspect", action="store_true", help="Read-only: download and dump the current layout")
    ap.add_argument("--force", action="store_true", help="Proceed even if the target month header already exists")
    ap.add_argument("--out", default=None, help="Where to write the built workbook (default /tmp/office_bills_*.xlsx)")
    args = ap.parse_args()

    print("◆ Loading credentials…")
    access_drive = drive_token()
    access_gmail, gmail_user = gmail_token()
    repo = repo_root()
    tmp = os.path.join("/tmp", "current_bills.xlsx")
    print(f"◆ Downloading master ({args.file_id})…")
    download_master(args.file_id, access_drive, tmp)

    if args.inspect:
        return cmd_inspect(args.file_id, access_drive, tmp)

    wb = openpyxl.load_workbook(tmp)
    ws = wb.active
    headers = find_month_headers(ws)
    if not headers:
        raise SystemExit("Could not find any 'Month YYYY' header in col A — run --inspect to see the layout.")

    # Resolve the target month. The master keeps the newest month at the TOP,
    # so headers[0] (the smallest row) is the most recent month to clone.
    prev_row, prev_label, (prev_year, prev_mon) = headers[0]
    if args.month:
        try:
            y, m = (int(x) for x in args.month.split("-"))
        except ValueError:
            raise SystemExit("--month must be YYYY-MM")
    else:
        y, m = (prev_year + 1, 1) if prev_mon == 12 else (prev_year, prev_mon + 1)
    target_label = f"{calendar.month_name[m]} {y}"
    target_ym = f"{y:04d}-{m:02d}"

    if any(label == target_label for _, label, _ in headers):
        if not args.force:
            raise SystemExit(f"⚠️  {target_label} already present in the master — aborting (use --force to override).")
        print(f"⚠️  {target_label} already present, but --force given — continuing.")

    print(f"◆ Previous month: {prev_label} (block starts row {prev_row})")
    print(f"◆ Target month:   {target_label} ({target_ym})")

    # Previous-month block = from its header row up to (but not including) the
    # next month header. Trailing blank rows are part of the month unit (each
    # month ends with spacer blanks before the next header) — clone them too so
    # the inter-month spacing is preserved exactly.
    next_header = next((r for r, label, _ in headers if r > prev_row), None)
    block_end = (next_header - 1) if next_header else ws.max_row
    n_rows = block_end - prev_row + 1
    print(f"◆ Cloning {n_rows} template rows (rows {prev_row}–{block_end}) to the top.")

    print(f"◆ Fetching {target_label} emails from {gmail_user}…")
    bills = fetch_bills(access_gmail, gmail_user, target_ym)

    # Insert the clone at the top, then copy the previous month's block into it.
    ws.insert_rows(1, amount=n_rows)
    clone_rows(ws, prev_row + n_rows, block_end + n_rows, 1)
    # New month label on the first row (the template's month header cell)
    ws.cell(1, 1).value = target_label

    updates = []
    print("◆ Overlaying extracted values…")
    overlay_bills(ws, bills, 1, n_rows, updates)
    for u in updates:
        print(u)

    print("◆ Validating…")
    top = find_month_headers(ws)
    assert top and top[0][1] == target_label, "validation failed: target month not at row 1"
    if top[0][0] != 1:
        print("  ⚠️  target month header not at row 1 (top[0] row =", top[0][0], ")")

    out_path = args.out or (f"/tmp/office_bills_{target_ym}{'_dryrun' if args.dry_run else ''}.xlsx")
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    print(f"◆ Wrote {out_path} ({os.path.getsize(out_path)} bytes)")

    if args.dry_run:
        print("\n🧪 DRY RUN — NOT uploaded. Review the file above, then re-run without --dry-run.")
        return 0

    print("◆ Uploading to Drive…")
    name = upload_master(args.file_id, access_drive, out_path)
    print(f"✅ Uploaded {name}")

    # Summary
    print("\n── Summary ──")
    total = 0.0
    for vendor in VENDOR_ORDER:
        for b in bills.get(vendor, []):
            if b.get("amount") is not None:
                total += b["amount"]
                print(f"  {vendor:<11} {b['account']:<14} {b['amount']:>10,.2f}  due {b.get('due_date') or ''}")
    print(f"  {'TOTAL':<26} {total:>10,.2f}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
